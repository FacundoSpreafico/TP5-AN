from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

def exportar_ejercicio1_excel(df, nombre_archivo='resultados_completos.xlsx'):
    try:
        if df.empty:
            raise ValueError("El DataFrame está vacío")

        required_cols = ['Imagen', 'Tiempo (s)', 'Centroide_x (µm)', 'Centroide_y (µm)',
                         'N_puntos_contorno', 'Contorno_x', 'Contorno_y']
        if not all(col in df.columns for col in required_cols):
            raise ValueError(f"DataFrame no contiene columnas requeridas: {required_cols}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Completos"

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        columnas = required_cols

        for col_num, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_style
            cell.fill = fill

        for row_num, row_data in enumerate(dataframe_to_rows(df[columnas], index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        column_widths = {
            'A': 20,  # Imagen
            'B': 12,  # Tiempo
            'C': 16,  # Centroide X
            'D': 16,  # Centroide Y
            'E': 18,  # N_puntos
            'F': 30,  # Contorno_x
            'G': 30  # Contorno_y
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                if cell.column_letter == 'B':
                    cell.number_format = '0.00000'
                elif cell.column_letter in ['C', 'D', 'E']:
                    cell.number_format = '0.00'

        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar a Excel: {str(e)}")
        return False


def exportar_ejercicio2_excel(df, nombre_archivo='resultados_completos2.xlsx'):
    try:
        if df.empty:
            raise ValueError("El DataFrame está vacío")

        wb = Workbook()
        ws = wb.active
        ws.title = "Ángulos de Contacto"

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Columnas específicas del Ejercicio 2
        columnas = [
            'Imagen', 'Imagen_num', 'Tiempo (s)', 'Angulo_izq', 'Angulo_der',
            'Perimetro_izq', 'Perimetro_der', 'Asimetria_perimetro',
            'Factor_esparcimiento', 'Tipo_angulo',
            'Centroide_x (µm)', 'Centroide_y (µm)', 'Centroide_estable'
        ]

        columnas_existentes = [col for col in columnas if col in df.columns]

        # Escribir encabezados
        for col_num, col_name in enumerate(columnas_existentes, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_style
            cell.fill = fill

        # Escribir filas
        for row_num, row_data in enumerate(dataframe_to_rows(df[columnas_existentes], index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar anchos de columna por nombre
        width_by_name = {
            'Imagen': 20,
            'Imagen_num': 12,
            'Tiempo (s)': 12,
            'Angulo_izq': 15,
            'Angulo_der': 15,
            'Perimetro_izq': 15,
            'Perimetro_der': 15,
            'Asimetria_perimetro': 18,
            'Factor_esparcimiento': 18,
            'Tipo_angulo': 15,
            'Centroide_x (µm)': 16,
            'Centroide_y (µm)': 16,
            'Centroide_estable': 16
        }
        for i, col in enumerate(columnas_existentes):
            letter = chr(ord('A') + i)
            ws.column_dimensions[letter].width = width_by_name.get(col, 15)

        # Formato numérico básico
        for row in ws.iter_rows(min_row=2, max_col=len(columnas_existentes)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        # Formatos específicos por columna (si existen)
        if 'Tiempo (s)' in columnas_existentes:
            idx = columnas_existentes.index('Tiempo (s)') + 1
            for col in ws.iter_cols(min_row=2, min_col=idx, max_col=idx):
                for cell in col:
                    cell.number_format = '0.00000'
        for nombre in ['Angulo_izq', 'Angulo_der', 'Perimetro_izq', 'Perimetro_der', 'Asimetria_perimetro', 'Factor_esparcimiento']:
            if nombre in columnas_existentes:
                idx = columnas_existentes.index(nombre) + 1
                for col in ws.iter_cols(min_row=2, min_col=idx, max_col=idx):
                    for cell in col:
                        cell.number_format = '0.00'

        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar a Excel: {str(e)}")
        return False


def exportar_ejercicio3_excel(df, nombre_archivo='resultados_completos3.xlsx'):
    """
    Exporta los resultados del Ejercicio 3 a un archivo Excel con formato específico.
    """
    try:
        if df.empty:
            raise ValueError("El DataFrame está vacío")

        wb = Workbook()
        ws = wb.active
        ws.title = "Propiedades Geométricas"

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Columnas específicas del Ejercicio 3
        columnas = ['Imagen', 'Tiempo (s)', 'Perimetro_izq (m)', 'Perimetro_der (m)',
                    'Simetria', 'Energia_cinetica (J)', 'Velocidad (m/s)',
                    'Altura_maxima (m)', 'Diametro_base (m)', 'Factor_esparcimiento']

        # Verificar qué columnas existen realmente en el DataFrame
        columnas_existentes = [col for col in columnas if col in df.columns]

        for col_num, col_name in enumerate(columnas_existentes, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_style
            cell.fill = fill

        for row_num, row_data in enumerate(dataframe_to_rows(df[columnas_existentes], index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar anchos de columna
        column_widths = {
            'A': 20,  # Imagen
            'B': 12,  # Tiempo
            'C': 18,  # Perimetro_izq (m)
            'D': 18,  # Perimetro_der (m)
            'E': 12,  # Simetria
            'F': 20,  # Energia_cinetica (J)
            'G': 15,  # Velocidad (m/s)
            'H': 18,  # Altura_maxima (m)
            'I': 18,  # Diametro_base (m)
            'J': 18  # Factor_esparcimiento
        }

        for col_letter, width in column_widths.items():
            if col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'][:len(columnas_existentes)]:
                ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                if cell.column_letter == 'B':
                    cell.number_format = '0.00000'
                elif cell.column_letter in ['C', 'D', 'F', 'G', 'H', 'I']:
                    cell.number_format = '0.0000E+00'
                elif cell.column_letter == 'E':
                    cell.number_format = '0.00'
                elif cell.column_letter == 'J':
                    cell.number_format = '0.00'

        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar a Excel: {str(e)}")
        return False