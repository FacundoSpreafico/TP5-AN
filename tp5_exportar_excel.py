from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import numpy as np


def exportar_ejercicio1_excel(df, nombre_archivo='resultados_tp5_volumen_area.xlsx'):
    """
    Exporta los resultados del Ejercicio 1 del TP5 a Excel con formato
    """
    try:
        if df.empty:
            raise ValueError("El DataFrame está vacío")

        wb = Workbook()
        ws = wb.active
        ws.title = "Volumenes_Areas"

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        columnas = ['Imagen', 'Tiempo (s)',
                    'Volumen_spline_trapecio', 'Volumen_spline_simpson',
                    'Volumen_poly_trapecio', 'Volumen_poly_simpson',
                    'Area_spline_trapecio', 'Area_spline_simpson',
                    'Area_poly_trapecio', 'Area_poly_simpson']

        columnas_existentes = [col for col in columnas if col in df.columns]

        for col_num, col_name in enumerate(columnas_existentes, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_style
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, row_data in enumerate(dataframe_to_rows(df[columnas_existentes], index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')

        column_widths = {
            'A': 20,  # Imagen
            'B': 12,  # Tiempo (s)
            'C': 26,  # Volumen_spline_trapecio
            'D': 26,  # Volumen_spline_simpson
            'E': 26,  # Volumen_poly_trapecio
            'F': 26,  # Volumen_poly_simpson
            'G': 24,  # Area_spline_trapecio
            'H': 24,  # Area_spline_simpson
            'I': 24,  # Area_poly_trapecio
            'J': 24   # Area_poly_simpson
        }

        for col_letter, width in column_widths.items():
            if col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'][:len(columnas_existentes)]:
                ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'B':  # Tiempo
                    cell.number_format = '0.00000'
                elif cell.column_letter in ['C', 'D', 'E', 'F']:  # Volúmenes
                    cell.number_format = '0.0000E+00'
                elif cell.column_letter in ['G', 'H', 'I', 'J']:  # Áreas
                    cell.number_format = '0.0000E+00'

        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar Ejercicio 1 a Excel: {str(e)}")
        return False


def exportar_ejercicio2_excel(resultados, solucionador, nombre_archivo='resultados_tp5_dinamica.xlsx'):
    """
    Exporta los resultados del Ejercicio 2 del TP5 a Excel con formato
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Hoja 1: Taylor Orden 3
        if 'taylor' in resultados:
            ws_taylor = wb.create_sheet("Taylor_Orden3")
            datos_taylor = resultados['taylor']

            headers = ['Tiempo (s)', 'Altura (m)', 'Velocidad (m/s)']
            for col_num, header in enumerate(headers, 1):
                cell = ws_taylor.cell(row=1, column=col_num, value=header)
                cell.font = header_style
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

            # Datos
            for i in range(len(datos_taylor['tiempos'])):
                ws_taylor.cell(row=i + 2, column=1, value=datos_taylor['tiempos'][i]).number_format = '0.00000'
                ws_taylor.cell(row=i + 2, column=2, value=datos_taylor['alturas'][i]).number_format = '0.0000E+00'
                ws_taylor.cell(row=i + 2, column=3, value=datos_taylor['velocidades'][i]).number_format = '0.0000E+00'

            ws_taylor.column_dimensions['A'].width = 12
            ws_taylor.column_dimensions['B'].width = 18
            ws_taylor.column_dimensions['C'].width = 18

            for row in ws_taylor.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        # Hoja 2: Runge-Kutta 5-6
        if 'rk56' in resultados:
            ws_rk56 = wb.create_sheet("Runge_Kutta_56")
            datos_rk56 = resultados['rk56']

            headers = ['Tiempo (s)', 'Altura (m)', 'Velocidad (m/s)']
            for col_num, header in enumerate(headers, 1):
                cell = ws_rk56.cell(row=1, column=col_num, value=header)
                cell.font = header_style
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

            for i in range(len(datos_rk56['tiempos'])):
                ws_rk56.cell(row=i + 2, column=1, value=datos_rk56['tiempos'][i]).number_format = '0.00000'
                ws_rk56.cell(row=i + 2, column=2, value=datos_rk56['alturas'][i]).number_format = '0.0000E+00'
                ws_rk56.cell(row=i + 2, column=3, value=datos_rk56['velocidades'][i]).number_format = '0.0000E+00'

            ws_rk56.column_dimensions['A'].width = 12
            ws_rk56.column_dimensions['B'].width = 18
            ws_rk56.column_dimensions['C'].width = 18

            for row in ws_rk56.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        # Hoja 3: Adams-Bashforth-Moulton
        if 'adams' in resultados:
            ws_adams = wb.create_sheet("Adams_Bashforth_Moulton")
            datos_adams = resultados['adams']

            headers = ['Tiempo (s)', 'Altura (m)', 'Velocidad (m/s)']
            for col_num, header in enumerate(headers, 1):
                cell = ws_adams.cell(row=1, column=col_num, value=header)
                cell.font = header_style
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

            for i in range(len(datos_adams['tiempos'])):
                ws_adams.cell(row=i + 2, column=1, value=datos_adams['tiempos'][i]).number_format = '0.00000'
                ws_adams.cell(row=i + 2, column=2, value=datos_adams['alturas'][i]).number_format = '0.0000E+00'
                ws_adams.cell(row=i + 2, column=3, value=datos_adams['velocidades'][i]).number_format = '0.0000E+00'

            ws_adams.column_dimensions['A'].width = 12
            ws_adams.column_dimensions['B'].width = 18
            ws_adams.column_dimensions['C'].width = 18

            for row in ws_adams.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

        # Hoja 4: Resumen Comparativo
        ws_resumen = wb.create_sheet("Resumen_Comparativo")

        headers = ['Metodo', 'Evaluaciones_Funcion', 'Error_RMS (m)', 'dt_Promedio (s)', 'Tolerancia']
        for col_num, header in enumerate(headers, 1):
            cell = ws_resumen.cell(row=1, column=col_num, value=header)
            cell.font = header_style
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        resumen_data = []
        for metodo in ['taylor', 'rk56', 'adams']:
            if metodo in resultados:
                datos = resultados[metodo]
                alturas_interp = np.interp(solucionador.tiempos_exp, datos['tiempos'], datos['alturas'])
                error_rms = np.sqrt(np.mean((alturas_interp - solucionador.alturas_exp) ** 2))

                resumen_data.append({
                    'Metodo': metodo,
                    'Evaluaciones_Funcion': datos['coste'],
                    'Error_RMS (m)': error_rms,
                    'dt_Promedio (s)': datos.get('dt_promedio', 'N/A'),
                    'Tolerancia': datos.get('tol_usada', 'N/A')
                })

        for i, datos in enumerate(resumen_data, 2):
            ws_resumen.cell(row=i, column=1, value=datos['Metodo']).alignment = Alignment(horizontal='center')
            ws_resumen.cell(row=i, column=2, value=datos['Evaluaciones_Funcion']).alignment = Alignment(
                horizontal='center')
            ws_resumen.cell(row=i, column=3, value=datos['Error_RMS (m)']).number_format = '0.0000E+00'
            ws_resumen.cell(row=i, column=3).alignment = Alignment(horizontal='center')

            if datos['dt_Promedio (s)'] != 'N/A':
                ws_resumen.cell(row=i, column=4, value=datos['dt_Promedio (s)']).number_format = '0.0000E+00'
            else:
                ws_resumen.cell(row=i, column=4, value=datos['dt_Promedio (s)'])
            ws_resumen.cell(row=i, column=4).alignment = Alignment(horizontal='center')

            if datos['Tolerancia'] != 'N/A':
                ws_resumen.cell(row=i, column=5, value=datos['Tolerancia']).number_format = '0.0000E+00'
            else:
                ws_resumen.cell(row=i, column=5, value=datos['Tolerancia'])
            ws_resumen.cell(row=i, column=5).alignment = Alignment(horizontal='center')

        ws_resumen.column_dimensions['A'].width = 20
        ws_resumen.column_dimensions['B'].width = 18
        ws_resumen.column_dimensions['C'].width = 15
        ws_resumen.column_dimensions['D'].width = 15
        ws_resumen.column_dimensions['E'].width = 15

        # Hoja 5: Parámetros del Modelo
        ws_parametros = wb.create_sheet("Parametros_Modelo")

        parametros = [
            ['Parámetro', 'Valor', 'Unidades'],
            ['Masa (m)', solucionador.m, 'kg'],
            ['Rigidez (k)', solucionador.k, 'N/m'],
            ['Amortiguamiento (c)', solucionador.c, 'Ns/m'],
            ['Altura equilibrio (yeq)', solucionador.yeq, 'm'],
            ['Tiempo inicial', solucionador.tiempos_exp[0], 's'],
            ['Tiempo final', solucionador.tiempos_exp[-1], 's']
        ]

        for row_num, fila in enumerate(parametros, 1):
            for col_num, valor in enumerate(fila, 1):
                cell = ws_parametros.cell(row=row_num, column=col_num, value=valor)
                if row_num == 1:
                    cell.font = header_style
                    cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
                if col_num == 2 and row_num > 1:  # Columna de valores
                    if isinstance(valor, float):
                        cell.number_format = '0.0000E+00'

        # Ajustar anchos de parámetros
        ws_parametros.column_dimensions['A'].width = 25
        ws_parametros.column_dimensions['B'].width = 18
        ws_parametros.column_dimensions['C'].width = 12

        wb.save(nombre_archivo)
        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar Ejercicio 2 a Excel: {str(e)}")
        return False